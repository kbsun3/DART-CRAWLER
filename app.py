import os
import time
import zipfile
import json
import xml.etree.ElementTree as ET
import datetime

import streamlit as st
import requests
import pandas as pd
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_OWNER_API_KEY  = (
    os.environ.get("DART_API_KEY")
    or st.secrets.get("DART_API_KEY", "")
)
_MASTER_CODE = (
    os.environ.get("MASTER_CODE")
    or st.secrets.get("MASTER_CODE", "")
)

def get_api_key() -> str:
    """사용할 DART API 키 반환.
    우선순위: 마스터코드 인증 → 사용자 직접 입력 키
    """
    if st.session_state.get("_master_auth"):
        return _OWNER_API_KEY
    return st.session_state.get("_user_api_key", "")
BASE_URL = "https://opendart.fss.or.kr/api"

REPORT_TYPES = {
    "연간 (사업보고서)": "11011",
    "반기 (반기보고서)": "11012",
    "1분기 (분기보고서)": "11013",
    "3분기 (분기보고서)": "11014",
}

FS_LABELS = {
    "BS":  "재무상태표",
    "IS":  "손익계산서",
    "CIS": "포괄손익계산서",
    "CF":  "현금흐름표",
    # SCE(자본변동표)는 2D 매트릭스 구조로 온전한 재현 불가 → L2 참조
}

THROTTLE_SEC = 0.2

# IFRS XBRL 표준 명명 규칙에서 현금유출을 양수로 저장하는 패턴
# 특정 회사·계정명이 아닌 IFRS full taxonomy 원칙 기반
_CF_OUTFLOW_PATTERNS = [
    "PurchaseOf",         # 자산 취득 (유형자산, 무형자산, 금융자산 등)
    "Repayments",         # 차입금·사채 상환
    "DividendsPaid",      # 배당금 지급
    "InterestPaid",       # 이자 지급
    "IncomeTaxesPaid",    # 법인세 납부
    "PaymentsFor",        # 리스부채 상환 등 (IFRS 16)
    "PaymentsTo",         # 직접법 영업CF (공급자/종업원 지급액)
    "AcquisitionOf",      # 자기주식 취득 등 dart_ 커스텀 취득 요소
]

# 위 패턴에 해당하지만 실제로는 이미 서명된(signed) 행을 보호하는 제외 패턴
_CF_OUTFLOW_EXCLUDE = [
    "CashFlowsFromUsedIn",   # 소계행 — DART가 이미 부호 처리
    "ProceedsFrom",           # 처분·상환 수취액(유입)
]


def _is_cf_outflow(account_id: str) -> bool:
    """
    IFRS XBRL 표준에서 현금유출을 양수로 저장하는 요소인지 판정.
    - ifrs-full_ 및 dart_ 커스텀 요소 모두 포함
    - 이미 서명된 소계행·유입 요소는 제외
    """
    if not account_id or "표준계정코드 미사용" in account_id:
        return False
    if any(exc in account_id for exc in _CF_OUTFLOW_EXCLUDE):
        return False
    if any(p in account_id for p in _CF_OUTFLOW_PATTERNS):
        return True
    # CashFlowsUsedIn 단독 (사업결합 순현금유출 등) — FromUsedIn 소계는 위에서 제외됨
    if "CashFlowsUsedIn" in account_id:
        return True
    return False


# 계정명 중복 레이블 매핑: account_id 키워드 → 표시 prefix
# 순서 중요: 먼저 매칭된 규칙이 적용됨
_DUPLICATE_LABEL_RULES: list[tuple[str, str]] = [
    # CIS 재분류 구분
    ("WillNotBeReclassifiedToProfitOrLoss", "[재분류X]"),
    ("WillBeReclassifiedToProfitOrLoss",    "[재분류O]"),
    # 유동/비유동 구분 (BS에서 가장 흔한 중복 원인)
    ("Shortterm",    "[단기]"),
    ("Current",      "[유동]"),
    ("Longterm",     "[장기]"),
    ("Noncurrent",   "[비유동]"),
]


def _make_display_nm(account_nm: str, account_id: str, is_dup: bool) -> str:
    """
    중복 계정명(is_dup=True)이면 account_id 기반 prefix를 붙여 구분.
    - 알려진 패턴은 의미 있는 한글 label 사용
    - 미등록 패턴은 account_id 말미를 fallback으로 사용
    """
    if not is_dup:
        return account_nm
    for keyword, label in _DUPLICATE_LABEL_RULES:
        if keyword in account_id:
            return f"{label} {account_nm}"
    # 비표준 커스텀 항목은 (기타) 표기
    if not account_id or "표준계정코드 미사용" in account_id:
        return f"{account_nm} (기타)"
    # 그 외 미등록 패턴은 account_id 말미로 구분
    suffix = account_id.split("_")[-1][:20]
    return f"{account_nm} [{suffix}]"


# ── Excel 스타일 상수 (디자인 사양 고정) ─────────────────────────────────────

# 배경색
_F_BLACK    = PatternFill("solid", fgColor="000000")  # 제목 배경
_F_NAVY     = PatternFill("solid", fgColor="1F4E78")  # 1차 헤더
_F_LBLUE    = PatternFill("solid", fgColor="D9E1F2")  # 2차 헤더 (결산일)
_F_YELLOW   = PatternFill("solid", fgColor="FFF2CC")  # 대분류 헤더
_F_GREEN    = PatternFill("solid", fgColor="E2EFDA")  # 합계/총계 (연한 초록)
_F_DK_GREEN = PatternFill("solid", fgColor="C6E0B4")  # 최종 결과 강조 (진한 초록)
_F_WHITE    = PatternFill("solid", fgColor="FFFFFF")  # 기본 배경

# 테두리
_S_THIN   = Side(style="thin",   color="000000")
_S_THICK  = Side(style="thick",  color="000000")
_S_NONE   = Side(style=None)
_S_DOUBLE = Side(style="double", color="000000")

# 글꼴
_FT_TITLE    = Font(name="맑은 고딕", bold=True,   color="FFFFFF", size=14)
_FT_H1       = Font(name="맑은 고딕", bold=True,   color="FFFFFF", size=11)
_FT_H2       = Font(name="맑은 고딕", bold=True,   color="000000", size=11)
_FT_BOLD     = Font(name="맑은 고딕", bold=True,   color="000000", size=11)
_FT_NORM     = Font(name="맑은 고딕",               color="000000", size=11)
_FT_ITALIC   = Font(name="맑은 고딕", italic=True,  color="000000", size=11)
_FT_UNIT     = Font(name="맑은 고딕", italic=True,  color="000000", size=9)
_FT_GREY     = Font(name="맑은 고딕",               color="BFBFBF", size=11)  # 빈값
_FT_GREY_NRM = Font(name="맑은 고딕",               color="7F7F7F", size=11)  # 보조 텍스트
_FT_GREY_BLD = Font(name="맑은 고딕", bold=True,    color="7F7F7F", size=11)  # 카테고리 라벨
_FT_GREY_ITA = Font(name="맑은 고딕", italic=True,  color="7F7F7F", size=11)  # % 보조지표
_FT_BLUE     = Font(name="맑은 고딕",               color="0000FF", size=11)  # 수동입력 강조
_FT_HIDDEN   = Font(color="FFFFFF")                                          # 숨김 행 (흰색 폰트)

# 정렬
_AL_C  = Alignment(horizontal="center", vertical="center")
_AL_L  = Alignment(horizontal="left",   vertical="center")
_AL_L1 = Alignment(horizontal="left",   vertical="center", indent=1)
_AL_L2 = Alignment(horizontal="left",   vertical="center", indent=2)
_AL_L3 = Alignment(horizontal="left",   vertical="center", indent=3)
_AL_R  = Alignment(horizontal="right",  vertical="center")

# 숫자 서식: 양수=콤마 / 음수=괄호 / 0=하이픈
_NUM_FMT = '#,##0;(#,##0);-'

# ── 하이라이트 레벨 (IFRS element name 기반) ─────────────────────────────────
# account_id의 마지막 _ 이후 element name으로 판정 (한글 계정명 의존 X)

_HL3_ELS = {    # Level 3: 합계/총계 → 초록 배경 + 굵게
    # BS
    "Assets", "EquityAndLiabilities", "LiabilitiesAndEquity",
    # IS / CIS (당기순이익, 포괄이익은 해당 재무제표에서만 최상위 합계)
    "ProfitLoss",
    "ComprehensiveIncome",
    # CF
    "CashAndCashEquivalentsAtEndOfPeriodCf",
}
_CF_ACTIVITY_ELS = {   # CF 3대 활동 소계 → Level 1 (노란색 굵게)
    "CashFlowsFromUsedInOperatingActivities",
    "CashFlowsFromUsedInInvestingActivities",
    "CashFlowsFromUsedInFinancingActivities",
}

# CF에서는 ProfitLoss·ComprehensiveIncome이 간접법 출발 세부항목
# (IS/CIS에서는 최상위 합계지만 CF에서는 Level 0이어야 함)
_CF_NOT_TOTAL = {"ProfitLoss", "ComprehensiveIncome"}


def _hl_level(account_id: str, has_values: bool, sj_div: str = "") -> int:
    """
    0 = 세부항목  (흰색, 보통, indent 3)
    1 = 대분류    (노랑, 굵게, 값 없음)  ― BS에서는 미사용
    2 = 중분류/소계 (흰색, 굵게, indent 1) ― BS에서는 미사용
    3 = 합계/총계  (초록, 굵게, indent 0)

    sj_div: 재무제표 구분 코드 (BS/IS/CIS/CF/SCE).
    """
    # ── BS ───────────────────────────────────────────────────────────────────
    if sj_div == "BS":
        if not account_id or "표준계정코드 미사용" in account_id:
            return 0
        el = account_id.rsplit("_", 1)[-1]
        if el in _HL3_ELS:
            return 3
        if el in {"Liabilities", "Equity"}:      # 부채총계, 자본총계 → 노란색
            return 1
        if el in {"CurrentAssets", "NoncurrentAssets"}:  # 유동/비유동자산 → 볼드
            return 2
        return 0

    # ── IS ───────────────────────────────────────────────────────────────────
    if sj_div == "IS":
        if not has_values:
            return 1  # 구조 헤더(당기순이익의 귀속 등) → indent 0
        if not account_id or "표준계정코드 미사용" in account_id:
            return 0
        el = account_id.rsplit("_", 1)[-1]
        if el in {"Revenue", "RevenueFromContractsWithCustomers"}:
            return 3  # 매출액 → 초록
        if el in {"ProfitLossFromOperatingActivities", "OperatingIncomeLoss",
                  "ProfitLoss"}:
            return 1  # 영업이익·당기순이익 → 노란색
        return 0

    # ── CIS ──────────────────────────────────────────────────────────────────
    if sj_div == "CIS":
        if not has_values:
            return 1  # 구조 헤더
        if not account_id or "표준계정코드 미사용" in account_id:
            return 0
        el = account_id.rsplit("_", 1)[-1]
        if el in {"ComprehensiveIncome"}:
            return 3  # 총포괄이익 → 초록
        if el in {"Revenue", "RevenueFromContractsWithCustomers"}:
            return 3  # 매출액 → 초록 (IS와 동일)
        if el in {"ProfitLossFromOperatingActivities", "OperatingIncomeLoss"}:
            return 1  # 영업이익 → 노란색 (IS와 동일)
        if el == "ProfitLoss":
            return 1  # 당기순이익 → 노란색
        if el == "OtherComprehensiveIncome":
            return 2  # 기타포괄손익 → 볼드
        return 0

    # ── SCE ──────────────────────────────────────────────────────────────────
    if sj_div == "SCE":
        if not has_values:
            return 1  # 구조 헤더
        if not account_id or "표준계정코드 미사용" in account_id:
            return 0
        el = account_id.rsplit("_", 1)[-1]
        # 기말잔액 계열 → 초록
        if "EndOfPeriod" in el or "ClosingBalance" in el:
            return 3
        # 기초잔액 계열 → 노란색
        if "BeginningOfPeriod" in el or "OpeningBalance" in el:
            return 1
        return 0

    # ── CF ───────────────────────────────────────────────────────────────────
    if not has_values:
        return 1  # 값 없는 행 = 대분류 헤더
    if not account_id or "표준계정코드 미사용" in account_id:
        return 0
    el = account_id.rsplit("_", 1)[-1]
    if el in _CF_NOT_TOTAL:
        return 0
    if el in _HL3_ELS:
        return 3
    if el in _CF_ACTIVITY_ELS:
        return 1  # 노란색 굵게
    return 0


def _write_fs_sheet(ws, pivot: pd.DataFrame, id_map: dict,
                    corp_name: str, fs_name: str,
                    years: list[int], period_map: dict,
                    fs_code: str = "") -> tuple[dict, list]:
    """재무제표 시트 작성 — 디자인 사양 준수.
    Returns: ({display_nm: excel_row}, year_cols)
      year_cols: FS 시트에 실제 존재하는 연도 목록 (CFO 참조 수식 컬럼 계산용)
    """
    year_cols = [c for c in pivot.columns if isinstance(c, int)]
    n_yr = len(year_cols)
    CB   = 2   # B열 = column index 2
    LAST = CB + n_yr  # 마지막 데이터 열

    def _w(row, col, value=None, fill=_F_WHITE, font=_FT_NORM,
           align=_AL_L, num_fmt=None, border=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.fill = fill; cell.font = font; cell.alignment = align
        if num_fmt:
            cell.number_format = num_fmt
        if border:
            cell.border = border
        return cell

    # ── 1행: 상단 여백 ────────────────────────────────────────────────────
    for c in range(CB, LAST + 1):
        _w(1, c)
    ws.row_dimensions[1].height = 8

    # ── 2행: 메인 제목 ────────────────────────────────────────────────────
    _w(2, CB, f"{corp_name} - {fs_name}", _F_BLACK, _FT_TITLE, _AL_L)
    for c in range(CB + 1, LAST + 1):
        _w(2, c, fill=_F_BLACK)
    ws.merge_cells(start_row=2, start_column=CB, end_row=2, end_column=LAST)
    ws.row_dimensions[2].height = 30

    # ── 3행: 단위 ─────────────────────────────────────────────────────────
    _w(3, CB, "(단위: 억원)", _F_WHITE, _FT_UNIT, _AL_L)
    for c in range(CB + 1, LAST + 1):
        _w(3, c)
    ws.row_dimensions[3].height = 16

    # ── 4행: 빈 줄 ────────────────────────────────────────────────────────
    for c in range(CB, LAST + 1):
        _w(4, c)
    ws.row_dimensions[4].height = 8

    # ── 5행: 1차 헤더 (FY 연도) ───────────────────────────────────────────
    _w(5, CB, "과  목", _F_NAVY, _FT_H1, _AL_C)
    for j, yr in enumerate(year_cols):
        _w(5, CB + 1 + j, f"FY{yr}", _F_NAVY, _FT_H1, _AL_C)
    ws.row_dimensions[5].height = 22

    # ── 6행: 2차 헤더 (기간 상세) ─────────────────────────────────────────
    _w(6, CB, "기간", _F_LBLUE, _FT_H2, _AL_C)
    for j, yr in enumerate(year_cols):
        _w(6, CB + 1 + j, period_map.get(yr, ""), _F_LBLUE, _FT_H2, _AL_C)
    ws.row_dimensions[6].height = 18

    # ── 7행: 빈 줄 ────────────────────────────────────────────────────────
    for c in range(CB, LAST + 1):
        _w(7, c)
    ws.row_dimensions[7].height = 6

    # ── 8행~: 데이터 (1패스: 셀 작성 + 레벨 기록) ────────────────────────
    DATA_START = 8
    row_levels: list[tuple[int, int]] = []  # (excel_row, level)
    nm_to_row: dict[str, int] = {}          # display_nm → excel_row (CFO 참조용)

    for i, (_, row_s) in enumerate(pivot.iterrows()):
        er = DATA_START + i
        display_nm = row_s["계정명"]
        nm_to_row[display_nm] = er
        account_id = id_map.get(display_nm, "")
        vals       = [row_s.get(yr) for yr in year_cols]
        has_vals   = any(v is not None and not pd.isna(v) for v in vals)
        level      = _hl_level(account_id, has_vals, fs_code)
        row_levels.append((er, level))

        # 행 스타일
        # Level 3(총계): indent 0 — 가장 돌출되게
        # Level 2(소계): indent 1
        # Level 1(헤더): indent 0
        # Level 0(세부): BS는 indent 1, 그 외 indent 3
        if level == 3:
            fill, font, nm_align = _F_GREEN,  _FT_BOLD, _AL_L
        elif level == 2:
            # BS: 부채총계/자본총계는 총계와 같은 indent 0 (볼드만 구분)
            # IS/CIS/CF: 소계는 indent 1
            sub_align = _AL_L if fs_code == "BS" else _AL_L1
            fill, font, nm_align = _F_WHITE,  _FT_BOLD, sub_align
        elif level == 1:
            fill, font, nm_align = _F_YELLOW, _FT_BOLD, _AL_L
        else:
            # BS: indent 1 / 그 외(IS·CIS·CF·SCE): indent 2
            detail_align = _AL_L1 if fs_code == "BS" else _AL_L2
            fill, font, nm_align = _F_WHITE,  _FT_NORM, detail_align

        _w(er, CB, display_nm, fill, font, nm_align)

        for j, yr in enumerate(year_cols):
            val  = row_s.get(yr)
            empty = val is None or pd.isna(val)
            fnt  = _FT_GREY if empty else font
            cell = _w(er, CB + 1 + j, None if empty else val,
                      fill, fnt, _AL_R, _NUM_FMT)

        ws.row_dimensions[er].height = 17

    # ── 2패스: 테두리 ─────────────────────────────────────────────────────
    # 마지막 총계 행 → double bottom
    last_total = max((r for r, lv in row_levels if lv == 3), default=None)

    for er, level in row_levels:
        if level in (2, 3):
            bot = _S_DOUBLE if er == last_total else _S_THIN
            bdr = Border(top=_S_THIN, bottom=bot)
            for c in range(CB, LAST + 1):
                ws.cell(row=er, column=c).border = bdr

    # ── 열 너비 / 틀 고정 ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[get_column_letter(CB)].width = 40
    for j in range(n_yr):
        ws.column_dimensions[get_column_letter(CB + 1 + j)].width = 17

    ws.freeze_panes = f"{get_column_letter(CB + 1)}8"
    return nm_to_row, year_cols


# ── Cash Flow Overview 분석 시트 ──────────────────────────────────────────────

def _cfo_extract(raw: pd.DataFrame, sj_div: str, patterns: list,
                 year_cols: list,
                 nm_fallbacks: list | None = None,
                 sj_div_fallback: str | None = None) -> tuple[dict, dict]:
    """raw에서 특정 XBRL 패턴의 계정을 연도별로 추출.
    Returns: (values_dict, display_nm_dict)
      values_dict    : {yr: value_or_None}
      display_nm_dict: {yr: display_nm_or_None}  ← FS 시트 참조 수식 생성용

    Fallback 순서:
      1) 당해 연도 행의 thstrm_amount (account_id 패턴 → account_nm 패턴)
      2) 다음 연도 행의 frmtrm_amount (동일 패턴 순서) ← 전기 비교컬럼 역이용
      3) sj_div_fallback 시트 동일 패턴 (e.g. IS 없으면 CIS 시도)
    """
    # sj_div_fallback: 일부 회사는 IS 없이 CIS만 사용 (포괄손익만 제출)
    divs_to_try = [sj_div]
    if sj_div_fallback:
        divs_to_try.append(sj_div_fallback)
    sub = raw[raw["sj_div"].isin(divs_to_try)]

    result:   dict = {yr: None for yr in year_cols}
    found_dnm: dict = {yr: None for yr in year_cols}

    def _search(rows: pd.DataFrame, amount_col: str):
        """rows에서 패턴 매칭 후 (value, display_nm) 반환, 없으면 (None, None)."""
        if amount_col not in rows.columns:
            return None, None
        # 1차: account_id 패턴
        if "account_id" in rows.columns:
            for pat in patterns:
                m = rows[rows["account_id"].str.contains(pat, na=False, regex=False)]
                if not m.empty:
                    v = m.iloc[0].get(amount_col)
                    if v is not None and not pd.isna(v):
                        dnm = m.iloc[0].get("display_nm")
                        return v, dnm
        # 2차: account_nm 한국어 fallback
        if nm_fallbacks and "account_nm" in rows.columns:
            for nm_pat in nm_fallbacks:
                m = rows[rows["account_nm"].str.contains(nm_pat, na=False, regex=False)]
                m = m[~m["account_nm"].str.contains("합계|총계", na=False)]
                if not m.empty:
                    v = m.iloc[0].get(amount_col)
                    if v is not None and not pd.isna(v):
                        dnm = m.iloc[0].get("display_nm")
                        return v, dnm
        return None, None

    for i, yr in enumerate(year_cols):
        # 1차: 당해 연도 thstrm
        result[yr], found_dnm[yr] = _search(sub[sub["year"] == yr], "thstrm_amount")
        # 2차: 다음 연도 frmtrm (전기 비교컬럼)
        if result[yr] is None and i + 1 < len(year_cols):
            nxt = year_cols[i + 1]
            v, dnm = _search(sub[sub["year"] == nxt], "frmtrm_amount")
            if v is not None:
                result[yr], found_dnm[yr] = v, dnm

    # ── 첫 연도 이전(n-1) 잔액 추출 ─────────────────────────────────────────
    # year_cols[0]의 frmtrm_amount = year_cols[0]-1 잔액
    # → NWC 첫 연도 증감 계산용 (출력 열 없이 내부 계산에만 사용)
    if year_cols:
        first_yr = year_cols[0]
        prior_yr = first_yr - 1
        if prior_yr not in result:
            v, _ = _search(sub[sub["year"] == first_yr], "frmtrm_amount")
            if v is not None:
                result[prior_yr] = v

    return result, found_dnm


def _cfo_add(a: dict, b: dict, year_cols: list) -> dict:
    """두 연도별 dict를 합산 (None 처리 포함)."""
    out = {}
    for yr in year_cols:
        va, vb = a.get(yr), b.get(yr)
        if va is None and vb is None:
            out[yr] = None
        else:
            out[yr] = (va or 0) + (vb or 0)
    return out


def _write_cfo_sheet(ws, raw: pd.DataFrame, corp_name: str,
                     year_cols: list, period_map: dict,
                     fs_row_maps: dict | None = None,
                     fs_sheet_names: dict | None = None,
                     fs_year_maps: dict | None = None) -> None:
    """Cash Flow Overview 분석 시트 작성.
    fs_row_maps   : {fs_code: {display_nm: excel_row}}  ← FS 시트 참조 수식용
    fs_year_maps  : {fs_code: [year, ...]}              ← FS 시트 연도→컬럼 매핑용
    fs_sheet_names: {fs_code: sheet_name_string}
    """

    PCT_FMT = "0.0%;(0.0%);-"
    CB  = 2   # B열 (label)
    NC  = len(year_cols)
    LAST = CB + NC  # 마지막 데이터 열

    def col(j):  # j=0 → C열, j=1 → D열 ...
        return get_column_letter(CB + 1 + j)

    def _w(row, c, val=None, fill=_F_WHITE, font=_FT_NORM,
           align=_AL_L, fmt=None, border=None):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill, cell.font, cell.alignment = fill, font, align
        if fmt:   cell.number_format = fmt
        if border: cell.border = border
        return cell

    def _make_border(top, bot):
        """top/bot Side 조합으로 Border 객체 반환. 둘 다 None이면 None."""
        if top and bot:
            return Border(top=top, bottom=bot)
        if top:
            return Border(top=top)
        if bot:
            return Border(bottom=bot)
        return None

    def _apply_label_border(r, top, bot):
        """label 열(CB)에 top/bot 테두리 적용."""
        if top or bot:
            ws.cell(row=r, column=CB).border = Border(
                top=top if top else _S_NONE,
                bottom=bot if bot else _S_NONE,
            )

    def _row(r, label, fill=_F_WHITE, font=_FT_NORM,
             vals=None, val_font=None, val_fill=None, fmt=_NUM_FMT,
             border_top=None, border_bot=None):
        """행 전체 작성 헬퍼."""
        _w(r, CB, label, fill, font, _AL_L)
        vf  = val_font or font
        vfl = val_fill or fill
        bdr = _make_border(border_top, border_bot)
        for j in range(NC):
            v = vals[j] if vals else None
            c = CB + 1 + j
            if v is None:
                _w(r, c, None, vfl, _FT_GREY, _AL_R, fmt, bdr)
            else:
                _w(r, c, v, vfl, vf, _AL_R, fmt, bdr)
        _apply_label_border(r, border_top, border_bot)
        ws.row_dimensions[r].height = 17

    def _formula_row(r, label, formula_tpl, fill=_F_WHITE, font=_FT_NORM,
                     fmt=_NUM_FMT, border_top=None, border_bot=None):
        """수식 행 작성 — formula_tpl에서 {c} 를 열 문자로 치환."""
        _w(r, CB, label, fill, font, _AL_L)
        bdr = _make_border(border_top, border_bot)
        for j in range(NC):
            c_ltr = col(j)
            formula = formula_tpl.format(c=c_ltr)
            _w(r, CB + 1 + j, formula, fill, font, _AL_R, fmt, bdr)
        _apply_label_border(r, border_top, border_bot)
        ws.row_dimensions[r].height = 17

    def _blank(r):
        for c in range(CB, LAST + 1):
            _w(r, c, None, _F_WHITE, _FT_NORM)
        ws.row_dimensions[r].height = 8

    # ── 데이터 추출 ──────────────────────────────────────────────────────────
    rev,  rev_dnm  = _cfo_extract(raw, "IS",
                                  ["Revenue", "RevenueFromContracts",
                                   "RevenueFromContractsWithCustomers"],
                                  year_cols,
                                  nm_fallbacks=["매출액", "영업수익"],
                                  sj_div_fallback="CIS")
    ebit, ebit_dnm = _cfo_extract(raw, "IS",
                                  ["ProfitLossFromOperatingActivities", "OperatingIncomeLoss"],
                                  year_cols,
                                  nm_fallbacks=["영업이익", "영업손익"],
                                  sj_div_fallback="CIS")
    cogs, _        = _cfo_extract(raw, "IS",
                                  ["CostOfSales", "CostOfGoodsSold"],
                                  year_cols,
                                  nm_fallbacks=["매출원가"],
                                  sj_div_fallback="CIS")
    ar,   ar_dnm   = _cfo_extract(raw, "BS",
                                  ["TradeAndOtherCurrentReceivables", "TradeReceivables",
                                   "TradeAndOtherReceivables", "CurrentTradeReceivables",
                                   "AccountsAndNotesReceivableCurrent"],
                                  year_cols, nm_fallbacks=["매출채권"])
    inv,  inv_dnm  = _cfo_extract(raw, "BS",
                                  ["Inventories", "CurrentInventories"],
                                  year_cols, nm_fallbacks=["재고자산"])
    ap,   ap_dnm   = _cfo_extract(raw, "BS",
                                  ["TradeAndOtherCurrentPayables", "TradeAndOtherPayables",
                                   "TradePayables", "CurrentTradePayables"],
                                  year_cols, nm_fallbacks=["매입채무"])
    ppe,   ppe_dnm   = _cfo_extract(raw, "CF",
                                    ["PurchaseOfPropertyPlantAndEquipment"], year_cols)
    intan, intan_dnm = _cfo_extract(raw, "CF",
                                    ["PurchaseOfIntangibleAssets"], year_cols)
    capex = _cfo_add(ppe, intan, year_cols)

    def yr_vals(d: dict) -> list:
        return [d.get(yr) for yr in year_cols]

    # ── FS 시트 참조 수식 헬퍼 ────────────────────────────────────────────────
    def _ref_vals_simple(values: dict, found_dnm: dict, fs_code: str,
                         fs_code_fallback: str | None = None) -> list:
        """FS 시트 셀 참조 수식. 없으면 값 직접 기입.

        핵심: FS 시트의 연도→컬럼 매핑(fs_year_maps)을 사용.
        CFO 시트와 FS 시트의 연도 집합이 달라도 올바른 컬럼 참조 생성.
        fs_code_fallback: 주 시트에서 행을 못 찾으면 fallback 시트도 시도 (IS→CIS 등).
        """
        dnm = next((v for v in found_dnm.values() if v is not None), None)

        def _resolve(code: str):
            """(sheet_name, row, fs_yr_cols) 또는 None 반환."""
            row_map  = (fs_row_maps  or {}).get(code, {})
            sheet_nm = (fs_sheet_names or {}).get(code, "")
            fs_yrs   = (fs_year_maps or {}).get(code, [])
            r = row_map.get(dnm) if (dnm and row_map) else None
            return (sheet_nm, r, fs_yrs) if r else None

        resolved = _resolve(fs_code)
        if resolved is None and fs_code_fallback:
            resolved = _resolve(fs_code_fallback)

        result = []
        for yr in year_cols:
            if resolved:
                sheet_nm, row, fs_yrs = resolved
                if yr in fs_yrs:
                    # 1순위: FS 시트 수식 참조 (재무제표와 연동)
                    c = get_column_letter(CB + 1 + fs_yrs.index(yr))
                    result.append(f"=IF('{sheet_nm}'!{c}{row}=\"\",\"\",'{sheet_nm}'!{c}{row})")
                    continue
            # 2순위: FS 행을 못 찾을 때만 Python 추출값 직접 기입
            py_val = values.get(yr)
            result.append(py_val)
        return result

    def _nwc_bal_vals(values: dict, found_dnm: dict,
                      fs_code: str, nm_keywords: list) -> list:
        """NWC Balance 행 값 생성.

        1순위: Python 추출값 (values dict)
        2순위: 정확한 display_nm 매칭으로 FS 시트 수식 참조 (_ref_vals_simple 방식)
        3순위: nm_keywords 키워드로 FS 시트 행 스캔 (완전 fallback)

        이 세 단계를 거쳐야 blank cell 없이 CCC 계산이 가능함.
        """
        row_map  = (fs_row_maps  or {}).get(fs_code, {})
        sheet_nm = (fs_sheet_names or {}).get(fs_code, "")
        fs_yrs   = (fs_year_maps or {}).get(fs_code, [])

        # 2순위: 정확한 display_nm 매칭
        dnm      = next((v for v in found_dnm.values() if v is not None), None)
        exact_r  = (row_map.get(dnm) if (dnm and row_map) else None)

        # 3순위: 키워드 기반 스캔 (2순위 실패 시)
        kw_r = None
        if exact_r is None and row_map:
            for nm, r in row_map.items():
                if any(kw in nm for kw in nm_keywords):
                    kw_r = r
                    break

        result = []
        for yr in year_cols:
            fs_r = exact_r if exact_r else kw_r
            if fs_r and yr in fs_yrs:
                # 1순위: FS 시트 수식 참조 (재무상태표와 연동)
                c = get_column_letter(CB + 1 + fs_yrs.index(yr))
                result.append(
                    f"=IF('{sheet_nm}'!{c}{fs_r}=\"\",\"\","
                    f"'{sheet_nm}'!{c}{fs_r})"
                )
            else:
                # 2순위: FS 행 못 찾을 때 Python 추출값
                result.append(values.get(yr))
        return result

    prior_yr = year_cols[0] - 1 if year_cols else None

    def _nwc_mv_formulas(bal_row: int, prior_row: int, outflow: bool) -> list:
        """NWC Movement: 모든 연도를 인시트 Excel 수식으로 생성.

        j=0 (첫 연도 N):  bal_row 현재열 vs prior_row 현재열
                          prior_row = N-1 잔액을 담은 숨김 행
        j>0 (이후 연도):  bal_row 현재열 vs bal_row 이전열

        outflow=True  → AR/Inv: -(cur-prv)  (증가=현금유출)
        outflow=False → AP:      cur-prv    (증가=현금유입)
        """
        result = []
        for j in range(len(year_cols)):
            c = get_column_letter(CB + 1 + j)
            if j == 0:
                # N-1 잔액은 prior_row의 같은 열(CB+1)에 있음
                pc = get_column_letter(CB + 1)
                if outflow:
                    f = (f"=IF(OR({c}{bal_row}=\"\",{pc}{prior_row}=\"\"),\"\","
                         f"-({c}{bal_row}-{pc}{prior_row}))")
                else:
                    f = (f"=IF(OR({c}{bal_row}=\"\",{pc}{prior_row}=\"\"),\"\","
                         f"{c}{bal_row}-{pc}{prior_row})")
            else:
                pc = get_column_letter(CB + j)
                if outflow:
                    f = (f"=IF(OR({c}{bal_row}=\"\",{pc}{bal_row}=\"\"),\"\","
                         f"-({c}{bal_row}-{pc}{bal_row}))")
                else:
                    f = (f"=IF(OR({c}{bal_row}=\"\",{pc}{bal_row}=\"\"),\"\","
                         f"{c}{bal_row}-{pc}{bal_row})")
            result.append(f)
        return result

    # ── 행 번호 정의 ─────────────────────────────────────────────────────────
    R = {
        "title": 2, "unit": 3, "blank1": 4,
        "hdr1": 5, "hdr2": 6, "blank2": 7,
        "rev": 8, "growth": 9,
        "ebit": 10, "da": 11, "ebitda": 12, "ebitda_pct": 13,
        "blank3": 14,
        "nwc_hdr": 15,
        "ar_chg": 16, "inv_chg": 17, "ap_chg": 18, "nwc_tot": 19,
        "blank4": 20,
        "ebitda_nwc": 21, "conv_pct": 22,
        "blank5": 23,
        "capex_hdr": 24,
        "capex_tot": 25, "capex_maint": 26, "capex_exp": 27,
        "capex_pct": 28,
        "blank6": 29,
        "fcf": 30, "cash_conv": 31,
        "blank7": 32,
        "ccc_hdr": 33,
        "nwc_bal_cat": 34,
        "ar_bal": 35, "inv_bal": 36, "ap_bal": 37,
        "blank8": 38,
        "ccc_cat": 39,
        "dso": 40, "dio": 41, "dpo": 42,
        "blank9": 43,
        "ccc": 44,
        # 숨김 행: N-1 잔액/매출 (첫 연도 수식 참조용, 시트에 표시 안 됨)
        "rev_prior":     46,
        "ar_bal_prior":  47,
        "inv_bal_prior": 48,
        "ap_bal_prior":  49,
        "cogs_hidden":   50,  # COGS 숨김 행 (DIO/DPO 수식 참조용)
        "ppe_hidden":    51,  # PPE 취득 숨김 행 (CF 참조)
        "intan_hidden":  52,  # 무형자산 취득 숨김 행 (CF 참조)
    }

    # ── 타이틀 ───────────────────────────────────────────────────────────────
    title_cell = ws.cell(row=R["title"], column=CB,
                         value=f"{corp_name} - Cash Flow Overview")
    title_cell.fill, title_cell.font, title_cell.alignment = (
        _F_BLACK, _FT_TITLE, _AL_L)
    for c in range(CB + 1, LAST + 1):
        _w(R["title"], c, fill=_F_BLACK, font=_FT_H1)
    ws.row_dimensions[R["title"]].height = 22

    _w(R["unit"], CB, "(단위: 억원)", _F_WHITE, _FT_UNIT, _AL_L)
    ws.row_dimensions[R["unit"]].height = 14
    _blank(R["blank1"])

    # ── 헤더 행 ──────────────────────────────────────────────────────────────
    _w(R["hdr1"], CB, "과 목", _F_NAVY, _FT_H1, _AL_C)
    for j, yr in enumerate(year_cols):
        _w(R["hdr1"], CB + 1 + j, f"FY{str(yr)[-2:]}", _F_NAVY, _FT_H1, _AL_C)
    ws.row_dimensions[R["hdr1"]].height = 18

    _w(R["hdr2"], CB, "기간", _F_LBLUE, _FT_H2, _AL_C)
    for j, yr in enumerate(year_cols):
        period = period_map.get(yr, "")
        _w(R["hdr2"], CB + 1 + j, period, _F_LBLUE, _FT_H2, _AL_C)
    ws.row_dimensions[R["hdr2"]].height = 18
    _blank(R["blank2"])

    # ── N-1 Revenue 숨김 행 (Growth % 첫 연도 수식 참조용) ──────────────────
    _rev_prior_val = rev.get(prior_yr)
    _row(R["rev_prior"], "", fill=_F_WHITE, font=_FT_HIDDEN,
         vals=[_rev_prior_val] + [None] * (NC - 1),
         val_font=_FT_HIDDEN)
    ws.row_dimensions[R["rev_prior"]].height = 0

    # ── Revenue ──────────────────────────────────────────────────────────────
    _row(R["rev"], "Revenue", font=_FT_BOLD,
         vals=_ref_vals_simple(rev, rev_dnm, "IS", fs_code_fallback="CIS"))
    # Growth %: 모든 연도를 인시트 Excel 수식으로 통일
    # 첫 연도: rev_prior 숨김 행 참조, 이후: 이전 열 참조
    _w(R["growth"], CB, "  Growth %", _F_WHITE, _FT_GREY_ITA, _AL_L)
    for j in range(NC):
        c_ltr = col(j)
        if j == 0:
            pc = get_column_letter(CB + 1)   # rev_prior 의 첫 번째 열
            formula = (f"=IF(AND({pc}{R['rev_prior']}<>\"\",{pc}{R['rev_prior']}<>0),"
                       f"{c_ltr}{R['rev']}/{pc}{R['rev_prior']}-1,\"\")")
        else:
            prev_c = col(j - 1)
            formula = (f"=IF(AND({prev_c}{R['rev']}<>\"\",{prev_c}{R['rev']}<>0),"
                       f"{c_ltr}{R['rev']}/{prev_c}{R['rev']}-1,\"\")")
        _w(R["growth"], CB + 1 + j, formula, _F_WHITE, _FT_GREY_ITA, _AL_R, PCT_FMT)
    ws.row_dimensions[R["growth"]].height = 17

    # ── EBIT / D&A / EBITDA ──────────────────────────────────────────────────
    _row(R["ebit"], "Operating Profit (EBIT)", font=_FT_BOLD,
         vals=_ref_vals_simple(ebit, ebit_dnm, "IS", fs_code_fallback="CIS"), border_bot=_S_THIN)
    # D&A — 수동 입력 (파란색 빈 셀)
    _w(R["da"], CB, "  (+) D&A", _F_WHITE, _FT_ITALIC, _AL_L)
    for j in range(NC):
        _w(R["da"], CB + 1 + j, None, _F_WHITE, _FT_BLUE, _AL_R, _NUM_FMT)
    ws.row_dimensions[R["da"]].height = 17

    _formula_row(R["ebitda"], "EBITDA",
                 f"=IF({{c}}{R['ebit']}=\"\",\"\",{{c}}{R['ebit']}+IF({{c}}{R['da']}=\"\",0,{{c}}{R['da']}))",
                 font=_FT_BOLD, border_top=_S_THIN)
    _formula_row(R["ebitda_pct"], "  EBITDA %",
                 f"=IF(AND({{c}}{R['rev']}<>\"\",{{c}}{R['rev']}<>0),{{c}}{R['ebitda']}/{{c}}{R['rev']},\"\")",
                 font=_FT_GREY_ITA, fmt=PCT_FMT)
    _blank(R["blank3"])

    # ── NWC Movement ─────────────────────────────────────────────────────────
    _w(R["nwc_hdr"], CB, "NWC Movement", _F_YELLOW, _FT_BOLD, _AL_L)
    for c in range(CB + 1, LAST + 1):
        _w(R["nwc_hdr"], c, fill=_F_YELLOW)
    ws.row_dimensions[R["nwc_hdr"]].height = 17

    # ── N-1 잔액 숨김 행 (NWC Movement 첫 연도 수식 참조용) ──────────────────
    # prior_yr 잔액을 첫 번째 데이터 열(CB+1)에만 기록. 나머지 열은 비워둠.
    for _r_key, _bal in [("ar_bal_prior", ar), ("inv_bal_prior", inv), ("ap_bal_prior", ap)]:
        _pv = _bal.get(prior_yr)
        _vals = [_pv] + [None] * (NC - 1)
        _row(R[_r_key], "", fill=_F_WHITE, font=_FT_HIDDEN,
             vals=_vals, val_font=_FT_HIDDEN)
        ws.row_dimensions[R[_r_key]].height = 0   # 완전 숨김

    # NWC Movement: 모든 연도를 인시트 Excel 수식으로 통일
    # (첫 연도도 N-1 숨김 행을 참조하므로 Python 계산 불필요)
    _row(R["ar_chg"],  "  매출채권 증감 (증가시 -)", font=_FT_NORM,
         vals=_nwc_mv_formulas(R["ar_bal"],  R["ar_bal_prior"],  outflow=True))
    _row(R["inv_chg"], "  재고자산 증감 (증가시 -)", font=_FT_NORM,
         vals=_nwc_mv_formulas(R["inv_bal"], R["inv_bal_prior"], outflow=True))
    _row(R["ap_chg"],  "  매입채무 증감 (증가시 +)", font=_FT_NORM,
         vals=_nwc_mv_formulas(R["ap_bal"],  R["ap_bal_prior"],  outflow=False),
         border_bot=_S_THIN)
    _formula_row(R["nwc_tot"], "Total NWC Movement",
                 f"=IF(AND({{c}}{R['ar_chg']}=\"\",{{c}}{R['inv_chg']}=\"\",{{c}}{R['ap_chg']}=\"\"),\"\","
                 f"IF({{c}}{R['ar_chg']}=\"\",0,{{c}}{R['ar_chg']})"
                 f"+IF({{c}}{R['inv_chg']}=\"\",0,{{c}}{R['inv_chg']})"
                 f"+IF({{c}}{R['ap_chg']}=\"\",0,{{c}}{R['ap_chg']}))",
                 font=_FT_BOLD, border_top=_S_THIN)
    _blank(R["blank4"])

    # ── EBITDA after NWC ─────────────────────────────────────────────────────
    _formula_row(R["ebitda_nwc"], "EBITDA after NWC Movement",
                 f"=IF({{c}}{R['ebitda']}=\"\",\"\",{{c}}{R['ebitda']}+IF({{c}}{R['nwc_tot']}=\"\",0,{{c}}{R['nwc_tot']}))",
                 fill=_F_GREEN, font=_FT_BOLD,
                 border_top=_S_THIN, border_bot=_S_THIN)
    _formula_row(R["conv_pct"], "  Conversion % (post NWC / EBITDA)",
                 f"=IF(AND({{c}}{R['ebitda']}<>\"\",{{c}}{R['ebitda']}<>0),{{c}}{R['ebitda_nwc']}/{{c}}{R['ebitda']},\"\")",
                 font=_FT_GREY_ITA, fmt=PCT_FMT)
    _blank(R["blank5"])

    # ── CAPEX ────────────────────────────────────────────────────────────────
    _w(R["capex_hdr"], CB, "CAPEX", _F_YELLOW, _FT_BOLD, _AL_L)
    for c in range(CB + 1, LAST + 1):
        _w(R["capex_hdr"], c, fill=_F_YELLOW)
    ws.row_dimensions[R["capex_hdr"]].height = 17

    # PPE·무형자산 숨김 행 → CF 시트 참조 (가능 시) or Python 값
    _row(R["ppe_hidden"],   "", fill=_F_WHITE, font=_FT_HIDDEN,
         vals=_ref_vals_simple(ppe,   ppe_dnm,   "CF"), val_font=_FT_HIDDEN)
    _row(R["intan_hidden"], "", fill=_F_WHITE, font=_FT_HIDDEN,
         vals=_ref_vals_simple(intan, intan_dnm, "CF"), val_font=_FT_HIDDEN)
    ws.row_dimensions[R["ppe_hidden"]].height   = 0
    ws.row_dimensions[R["intan_hidden"]].height = 0

    # Total CAPEX = PPE + 무형자산 (in-sheet 수식)
    _formula_row(R["capex_tot"], "  Total CAPEX",
                 f"=IF(AND({{c}}{R['ppe_hidden']}=\"\",{{c}}{R['intan_hidden']}=\"\"),\"\","
                 f"IF({{c}}{R['ppe_hidden']}=\"\",0,{{c}}{R['ppe_hidden']})"
                 f"+IF({{c}}{R['intan_hidden']}=\"\",0,{{c}}{R['intan_hidden']}))",
                 font=_FT_BOLD)
    # Maintenance / Expansion — 수동 입력 (파란색)
    _w(R["capex_maint"], CB, "    ㄴ Maintenance CAPEX", _F_WHITE, _FT_GREY_NRM, _AL_L)
    _w(R["capex_exp"],   CB, "    ㄴ Expansion CAPEX",   _F_WHITE, _FT_GREY_NRM, _AL_L)
    for j in range(NC):
        _w(R["capex_maint"], CB + 1 + j, None, _F_WHITE, _FT_BLUE, _AL_R, _NUM_FMT)
        _w(R["capex_exp"],   CB + 1 + j, None, _F_WHITE, _FT_BLUE, _AL_R, _NUM_FMT)
    ws.row_dimensions[R["capex_maint"]].height = 17
    ws.row_dimensions[R["capex_exp"]].height   = 17

    _formula_row(R["capex_pct"], "  CAPEX % of Revenue",
                 f"=IF(AND({{c}}{R['rev']}<>\"\",{{c}}{R['rev']}<>0),IF({{c}}{R['capex_tot']}=\"\",\"\",{{c}}{R['capex_tot']}/{{c}}{R['rev']}),\"\")",
                 font=_FT_GREY_ITA, fmt=PCT_FMT, border_bot=_S_THIN)
    _blank(R["blank6"])

    # ── Free Cash Flow ────────────────────────────────────────────────────────
    _formula_row(R["fcf"], "Free Cash Flow",
                 f"=IF({{c}}{R['ebitda_nwc']}=\"\",\"\",{{c}}{R['ebitda_nwc']}-IF({{c}}{R['capex_tot']}=\"\",0,{{c}}{R['capex_tot']}))",
                 fill=_F_DK_GREEN, font=_FT_BOLD,
                 border_top=_S_THICK, border_bot=_S_DOUBLE)
    _formula_row(R["cash_conv"], "  Cash Conversion % (FCF / EBITDA)",
                 f"=IF(AND({{c}}{R['ebitda']}<>\"\",{{c}}{R['ebitda']}<>0),{{c}}{R['fcf']}/{{c}}{R['ebitda']},\"\")",
                 font=_FT_GREY_ITA, fmt=PCT_FMT)
    _blank(R["blank7"])

    # ── NWC Balance & CCC ────────────────────────────────────────────────────
    _w(R["ccc_hdr"], CB, "NWC Balance & Cash Conversion Cycle (CCC)",
       _F_YELLOW, _FT_BOLD, _AL_L)
    for c in range(CB + 1, LAST + 1):
        _w(R["ccc_hdr"], c, fill=_F_YELLOW)
    ws.row_dimensions[R["ccc_hdr"]].height = 17

    _w(R["nwc_bal_cat"], CB, "[NWC Balance]", _F_WHITE, _FT_GREY_BLD, _AL_L)
    ws.row_dimensions[R["nwc_bal_cat"]].height = 17

    # NWC Balance: Python 값 → 정확한 display_nm → 키워드 스캔 순으로 fallback
    # _nwc_bal_vals 는 3단계 fallback으로 가장 robust한 방법
    _row(R["ar_bal"],  "  매출채권", font=_FT_NORM,
         vals=_nwc_bal_vals(ar,  ar_dnm,  "BS", ["매출채권", "receivable", "Receivable"]))
    _row(R["inv_bal"], "  재고자산", font=_FT_NORM,
         vals=_nwc_bal_vals(inv, inv_dnm, "BS", ["재고자산", "nventor", "Inventor"]))
    _row(R["ap_bal"],  "  매입채무", font=_FT_NORM,
         vals=_nwc_bal_vals(ap,  ap_dnm,  "BS", ["매입채무", "payable",  "Payable"]))
    _blank(R["blank8"])

    _w(R["ccc_cat"], CB, "[Cash Conversion Cycle (days)]", _F_WHITE, _FT_GREY_BLD, _AL_L)
    ws.row_dimensions[R["ccc_cat"]].height = 17

    # DSO / DIO / DPO 수식 (COGS 있으면 사용, 없으면 Revenue fallback)
    cogs_vals = yr_vals(cogs)
    has_cogs  = any(v is not None for v in cogs_vals)
    denom_row = R["rev"]   # fallback

    if has_cogs:
        # cogs를 별도 숨김 행에 기록 (DIO/DPO 수식 참조용)
        # 레이블·폰트 모두 흰색으로 완전 은폐
        cogs_row = R["cogs_hidden"]  # row 50 — rev_prior(46)~ap_bal_prior(49)와 충돌 방지
        _row(cogs_row, "", fill=_F_WHITE, font=_FT_HIDDEN, vals=cogs_vals,
             val_font=_FT_HIDDEN)
        ws.row_dimensions[cogs_row].height = 0  # 행 높이 0 (완전 숨김)
        denom_row = cogs_row

    # DSO/DIO/DPO: balance 셀도 빈값 체크 추가 → #VALUE! 방지
    _formula_row(R["dso"], "  DSO (매출채권 회수일수)",
                 f"=IF(AND({{c}}{R['rev']}<>\"\",{{c}}{R['rev']}<>0,"
                 f"{{c}}{R['ar_bal']}<>\"\"),{{c}}{R['ar_bal']}/{{c}}{R['rev']}*365,\"\")",
                 font=_FT_NORM, fmt="0.0")
    _formula_row(R["dio"], "  DIO (재고 회전일수)",
                 f"=IF(AND({{c}}{denom_row}<>\"\",{{c}}{denom_row}<>0,"
                 f"{{c}}{R['inv_bal']}<>\"\"),{{c}}{R['inv_bal']}/{{c}}{denom_row}*365,\"\")",
                 font=_FT_NORM, fmt="0.0")
    _formula_row(R["dpo"], "  DPO (매입채무 지급일수)",
                 f"=IF(AND({{c}}{denom_row}<>\"\",{{c}}{denom_row}<>0,"
                 f"{{c}}{R['ap_bal']}<>\"\"),{{c}}{R['ap_bal']}/{{c}}{denom_row}*365,\"\")",
                 font=_FT_NORM, fmt="0.0", border_bot=_S_THIN)
    _blank(R["blank9"])
    _formula_row(R["ccc"], "CCC (DSO + DIO - DPO)",
                 f"=IF({{c}}{R['dso']}=\"\",\"\",{{c}}{R['dso']}+{{c}}{R['dio']}-{{c}}{R['dpo']})",
                 fill=_F_GREEN, font=_FT_BOLD, fmt="0.0",
                 border_top=_S_THIN, border_bot=_S_DOUBLE)

    # ── 입력 가이드 (파란 입력 셀 바로 옆에만 표시) ──────────────────────────
    GC = LAST + 2   # 가이드 열 (데이터 마지막 열 + 여백 1)

    _GF_HDR  = PatternFill("solid", fgColor="1F4E78")
    _GF_IN   = PatternFill("solid", fgColor="EFF6FF")   # 입력 행 배경
    _GF_NONE = _F_WHITE

    _GT_HDR  = Font(name="맑은 고딕", bold=True,  color="FFFFFF", size=10)
    _GT_IN   = Font(name="맑은 고딕",              color="1E3A5F", size=10)
    _GT_SUB  = Font(name="맑은 고딕", italic=True, color="4B5563", size=9)
    _GT_AL     = Alignment(horizontal="left",   vertical="center",
                           wrap_text=False, indent=1)
    _GT_AL_CTR = Alignment(horizontal="center", vertical="center",
                           wrap_text=False)

    def _gw(r, text, fill=_GF_NONE, font=_GT_IN, height=None, align=None):
        cell = ws.cell(row=r, column=GC, value=text)
        cell.fill, cell.font, cell.alignment = fill, font, align or _GT_AL
        if height:
            ws.row_dimensions[r].height = height

    # 헤더 (타이틀 행) — 가운데 정렬
    _gw(R["title"], "[입력 가이드]  파란 셀 3개만 직접 입력",
        fill=_GF_HDR, font=_GT_HDR, align=_GT_AL_CTR)

    # ① D&A — 입력 행(row 11) 바로 옆
    _gw(R["da"],
        "① D&A: 감가상각 + 상각비 합산  (현금흐름표 영업CF 조정항목 참조)",
        fill=_GF_IN, font=_GT_IN, height=20)

    # ② Maintenance CAPEX — 입력 행(row 26) 바로 옆
    _gw(R["capex_maint"],
        "② Maintenance CAPEX: 기존 설비 유지·교체 투자금",
        fill=_GF_IN, font=_GT_IN, height=20)

    # ③ Expansion CAPEX — 입력 행(row 27) 바로 옆
    _gw(R["capex_exp"],
        "③ Expansion CAPEX: Total CAPEX − ②",
        fill=_GF_IN, font=_GT_IN, height=20)

    # 가이드 열 너비 — 내용에 맞게 조정
    ws.column_dimensions[get_column_letter(LAST + 1)].width = 1   # 여백
    ws.column_dimensions[get_column_letter(GC)].width = 80

    # ── 열 너비 / 틀 고정 ─────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions[get_column_letter(CB)].width = 44
    for j in range(NC):
        ws.column_dimensions[get_column_letter(CB + 1 + j)].width = 17
    ws.freeze_panes = f"{get_column_letter(CB + 1)}{R['hdr1']}"
    ws.sheet_view.showGridLines = False


# ── 데이터 로직 ───────────────────────────────────────────────────────────────

@st.cache_data(ttl=86400, show_spinner=False)
def load_corp_codes() -> pd.DataFrame:
    # 1순위: repo에 번들된 파일 (Streamlit Cloud에서 DART API 차단 대비)
    _here = os.path.dirname(os.path.abspath(__file__))
    _bundle = os.path.join(_here, "corp_codes.json")
    if os.path.exists(_bundle):
        with open(_bundle, encoding="utf-8") as f:
            return pd.DataFrame(json.load(f))
    # 2순위: DART API 직접 호출 (로컬 실행 또는 파일 없을 때)
    r = requests.get(f"{BASE_URL}/corpCode.xml", params={"crtfc_key": get_api_key()}, timeout=30)
    r.raise_for_status()
    z = zipfile.ZipFile(BytesIO(r.content))
    root = ET.fromstring(z.read("CORPCODE.xml"))
    rows = [
        {
            "corp_code":  item.findtext("corp_code", "").strip(),
            "corp_name":  item.findtext("corp_name", "").strip(),
            "stock_code": item.findtext("stock_code", "").strip(),
            "modify_date":item.findtext("modify_date", "").strip(),
        }
        for item in root.findall("list")
    ]
    return pd.DataFrame(rows)


def search_company(query: str, corp_df: pd.DataFrame) -> pd.DataFrame:
    query = query.strip()
    if not query:
        return pd.DataFrame()
    if query.isdigit() and len(query) == 6:
        result = corp_df[corp_df["stock_code"] == query]
    else:
        result = corp_df[corp_df["corp_name"].str.contains(query, na=False)]
    listed   = result[result["stock_code"] != ""].copy()
    unlisted = result[result["stock_code"] == ""].copy()
    return pd.concat([listed, unlisted], ignore_index=True)


@st.cache_data(ttl=86400, show_spinner=False)
def fetch_financials(corp_code: str, year: int, reprt_code: str,
                     api_key: str = "") -> pd.DataFrame:
    key = api_key or _DEFAULT_API_KEY
    for fs_div in ("CFS", "OFS"):
        time.sleep(THROTTLE_SEC)
        r = requests.get(
            f"{BASE_URL}/fnlttSinglAcntAll.json",
            params={"crtfc_key": key, "corp_code": corp_code,
                    "bsns_year": str(year), "reprt_code": reprt_code, "fs_div": fs_div},
            timeout=15,
        )
        data = r.json()
        if data.get("status") == "000" and data.get("list"):
            df = pd.DataFrame(data["list"])
            df["year"] = year
            df["fs_div_used"] = fs_div
            return df
    return pd.DataFrame()


def parse_amount(val):
    try:
        cleaned = str(val).replace(",", "").strip()
        return int(cleaned) if cleaned not in ("", "-", "None") else None
    except (ValueError, TypeError):
        return None


def build_excel(corp_code: str, corp_name: str, stock_code: str,
                years: list[int], reprt_code: str) -> bytes:
    all_frames = []
    progress = st.progress(0, text="데이터 수집 중...")

    for i, year in enumerate(years):
        progress.progress(i / len(years), text=f"{year}년 재무제표 수집 중...")
        df = fetch_financials(corp_code, year, reprt_code, api_key=get_api_key())
        if not df.empty:
            all_frames.append(df)

    progress.progress(1.0, text="수집 완료")

    if not all_frames:
        st.error("재무 데이터를 가져오지 못했습니다. 연도나 보고서 종류를 확인해주세요.")
        return b""

    raw = pd.concat(all_frames, ignore_index=True)

    fs_div_used = raw["fs_div_used"].iloc[0] if "fs_div_used" in raw.columns else "?"
    fs_label = "연결재무제표" if fs_div_used == "CFS" else "별도재무제표"

    for col in ["thstrm_amount", "frmtrm_amount", "bfefrmtrm_amount"]:
        if col in raw.columns:
            raw[col] = raw[col].apply(parse_amount)

    # Bug #2 수정: CF 세부 라인 부호 복원
    if "account_id" in raw.columns:
        cf_mask = raw["sj_div"] == "CF"
        outflow_mask = raw["account_id"].apply(
            lambda aid: _is_cf_outflow(str(aid) if aid else "")
        )
        for col in ["thstrm_amount", "frmtrm_amount", "bfefrmtrm_amount"]:
            if col in raw.columns:
                negate = cf_mask & outflow_mask & raw[col].notna() & (raw[col] > 0)
                raw.loc[negate, col] = -raw.loc[negate, col]

    # ── 단위 변환: 원 → 억원 (반올림) ─────────────────────────────────────────
    _EOKWON = 100_000_000
    for col in ["thstrm_amount", "frmtrm_amount", "bfefrmtrm_amount"]:
        if col in raw.columns:
            raw[col] = raw[col].apply(
                lambda v: round(v / _EOKWON) if v is not None and not pd.isna(v) else v
            )

    # Fix #3: 연도별 계정명 변경(리네임) 정규화
    # 동일 account_id가 연도마다 다른 account_nm을 가질 때 가장 최근 연도 기준으로 통일.
    # 예: '영업활동현금흐름표'(FY21) vs '영업활동현금흐름'(FY24) → 동일 IFRS 요소
    # 비표준 계정('-표준계정코드 미사용-')은 account_id가 없으므로 account_nm 유지.
    if "account_id" in raw.columns:
        non_std_mask = raw["account_id"].str.contains("표준계정코드 미사용", na=False)
        std_rows = raw[~non_std_mask]
        if not std_rows.empty:
            # account_id 별 최신 연도의 account_nm을 정규 명칭으로 채택
            canonical = (
                std_rows.sort_values("year", ascending=False)
                .groupby("account_id")["account_nm"]
                .first()
                .to_dict()
            )
            raw.loc[~non_std_mask, "account_nm"] = raw.loc[~non_std_mask, "account_id"].map(
                lambda aid: canonical.get(aid, "")
            )

    # Bug #1 수정: 재무제표·연도 내 중복 account_nm을 자동 감지 후 display_nm 생성
    if "account_id" in raw.columns:
        # 중복 여부: 동일 sj_div + year 안에서 account_nm이 2회 이상 등장하는 경우
        dup_keys = set(
            raw.groupby(["sj_div", "year", "account_nm"])
               .size()
               .loc[lambda s: s > 1]
               .reset_index()[["sj_div", "account_nm"]]
               .apply(tuple, axis=1)
        )
        raw["display_nm"] = raw.apply(
            lambda r: _make_display_nm(
                r.get("account_nm", ""),
                str(r.get("account_id", "")),
                is_dup=(r.get("sj_div"), r.get("account_nm")) in dup_keys,
            ),
            axis=1,
        )
    else:
        raw["display_nm"] = raw["account_nm"]

    # display_nm → account_id 매핑 (하이라이트 레벨 판정용)
    id_map = {}
    if "account_id" in raw.columns:
        id_map = raw.groupby("display_nm")["account_id"].first().to_dict()

    # year → 기간 문자열 (헤더 서브행용)
    period_map = {}
    if "thstrm_nm" in raw.columns:
        period_map = raw.groupby("year")["thstrm_nm"].first().to_dict()

    output = BytesIO()
    try:
      with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 정보 시트
        try:
            meta = pd.DataFrame([
                {"항목": "기업명",           "내용": corp_name},
                {"항목": "종목코드",         "내용": stock_code or "-"},
                {"항목": "DART corp_code",   "내용": corp_code},
                {"항목": "재무제표 종류",    "내용": fs_label},
                {"항목": "수집 연도",        "내용": f"{years[0]}~{years[-1]}"},
                {"항목": "수집일",           "내용": pd.Timestamp.now().strftime("%Y-%m-%d")},
            ])
            meta.to_excel(writer, sheet_name="정보", index=False)
        except Exception as _e:
            pass  # 정보 시트 실패해도 계속 진행

        # 원본데이터 시트
        try:
            cols_keep = [c for c in ["year", "sj_div", "sj_nm", "account_id",
                                      "account_nm", "thstrm_amount", "fs_div_used"]
                         if c in raw.columns]
            raw[cols_keep].rename(columns={
                "year": "연도", "sj_div": "재무제표구분코드", "sj_nm": "재무제표명",
                "account_id": "계정ID", "account_nm": "계정명",
                "thstrm_amount": "금액(억원)", "fs_div_used": "연결/별도",
            }).to_excel(writer, sheet_name="원본데이터", index=False)
        except Exception as _e:
            pass  # 원본데이터 시트 실패해도 계속 진행

        # 재무제표별 스타일 시트
        fs_row_maps:  dict[str, dict] = {}   # {fs_code: {display_nm: excel_row}}
        fs_year_maps: dict[str, list] = {}   # {fs_code: [year, ...]}  ← 컬럼 위치 계산용
        for fs_code, fs_name in FS_LABELS.items():
            subset = raw[raw["sj_div"] == fs_code].copy()
            if subset.empty:
                continue

            if "ord" in subset.columns:
                subset["ord"] = pd.to_numeric(subset["ord"], errors="coerce")
                subset = subset.sort_values(["year", "ord"], na_position="last")

            # SCE: 자본총계(Equity) 열만 추출 → 난잡한 2D 매트릭스 대신 단일 컬럼 뷰
            if fs_code == "SCE" and "account_id" in subset.columns:
                sce_total_mask = subset["account_id"].str.contains(
                    r"(?:^|_)Equity$", regex=True, na=False
                )
                if sce_total_mask.any():
                    subset = subset[sce_total_mask]

            pivot = (
                subset.groupby(["display_nm", "year"], sort=False)["thstrm_amount"]
                .first().unstack("year").reset_index()
            )
            pivot.columns.name = None
            pivot = pivot.rename(columns={"display_nm": "계정명"})
            year_cols = sorted([c for c in pivot.columns if isinstance(c, int)])
            pivot = pivot[["계정명"] + year_cols]

            # ── frmtrm / bfefrmtrm 백필 ──────────────────────────────────────
            # DART는 한 번의 쿼리에 당기(thstrm)·전기(frmtrm)·전전기(bfefrmtrm)를
            # 모두 반환함. thstrm 단독 pivot에서 생긴 NaN을 이미 받아온 값으로 보완.
            # 추정·예측 없음 — API 응답에 이미 있는 숫자만 사용.
            bf: dict[tuple, float] = {}
            for col, offset in [("frmtrm_amount", 1), ("bfefrmtrm_amount", 2)]:
                if col not in subset.columns:
                    continue
                for _, r in subset.iterrows():
                    val = r.get(col)
                    if val is None or pd.isna(val):
                        continue
                    target = int(r["year"]) - offset
                    if target not in year_cols:
                        continue
                    key = (r["display_nm"], target)
                    if key not in bf:          # 인접 연도(frmtrm) 우선
                        bf[key] = val

            if bf:
                for idx, row_p in pivot.iterrows():
                    dnm = row_p["계정명"]
                    for yr in year_cols:
                        if pd.isna(row_p.get(yr)):
                            bv = bf.get((dnm, yr))
                            if bv is not None:
                                pivot.at[idx, yr] = bv

            # 전 기간 NaN 행 제거 ─────────────────────────────────────────────
            # · 표준 account_id 있는 행 → NaN이어도 유지 (구조 헤더: 당기순이익의 귀속 등)
            # · "-표준계정코드 미사용-" 행 → NaN이면 제거 (inapplicable placeholder)
            has_val = pivot[year_cols].notna().any(axis=1)
            is_nonstandard = pivot["계정명"].map(
                lambda nm: "표준계정코드 미사용" in id_map.get(nm, "")
            )
            pivot = pivot[has_val | ~is_nonstandard].reset_index(drop=True)

            sheet_name = fs_name[:31]
            writer.book.create_sheet(sheet_name)
            ws = writer.book[sheet_name]
            try:
                nm_to_row, fs_yrs = _write_fs_sheet(ws, pivot, id_map, corp_name, fs_name,
                                                    year_cols, period_map, fs_code)
                fs_row_maps[fs_code]  = nm_to_row
                fs_year_maps[fs_code] = fs_yrs
            except Exception as _fs_err:
                ws.cell(row=2, column=2, value=f"[시트 생성 오류] {_fs_err}")
                fs_row_maps[fs_code]  = {}
                fs_year_maps[fs_code] = []

        # ── Cash Flow Overview 분석 시트 ──────────────────────────────────
        all_year_cols = sorted(raw["year"].unique().tolist())
        ws_cfo = writer.book.create_sheet("Cash Flow Overview")
        try:
            _write_cfo_sheet(ws_cfo, raw, corp_name, all_year_cols, period_map,
                             fs_row_maps=fs_row_maps,
                             fs_sheet_names={k: v for k, v in FS_LABELS.items()},
                             fs_year_maps=fs_year_maps)
        except Exception as _cfo_err:
            ws_cfo.cell(row=2, column=2, value=f"[CFO 시트 생성 오류] {_cfo_err}")

        # ── 시트 순서 조정: Cash Flow Overview → 원본데이터 바로 다음 ──────
        _wb = writer.book
        try:
            _cfo_ws = _wb["Cash Flow Overview"]
            _wb._sheets.remove(_cfo_ws)
            _raw_idx = _wb.sheetnames.index("원본데이터")
            _wb._sheets.insert(_raw_idx + 1, _cfo_ws)
        except Exception:
            pass  # 순서 조정 실패해도 시트 내용에는 영향 없음

    except Exception as _wb_err:
        # with 블록 자체의 예외 (ExcelWriter 초기화 실패 등)
        # → 빈 워크북 하나 만들어서 에러 메시지 기록 후 반환
        import traceback
        from openpyxl import Workbook
        wb = Workbook()
        ws_err = wb.active
        ws_err.title = "오류"
        ws_err["A1"] = "Excel 생성 중 오류 발생"
        ws_err["A2"] = str(_wb_err)
        ws_err["A3"] = traceback.format_exc()
        err_out = BytesIO()
        wb.save(err_out)
        err_out.seek(0)
        try:
            progress.empty()
        except Exception:
            pass
        return err_out.getvalue()

    # ── 정상 반환 ─────────────────────────────────────────────────────────────
    # finally 안에서 예외가 나면 return이 취소되므로 finally를 쓰지 않음.
    # progress 제거는 여기서 직접 처리.
    try:
        progress.empty()
    except Exception:
        pass

    # ZIP 매직바이트(PK\x03\x04) 검증 — BytesIO가 비거나 손상됐을 경우 fallback
    output.seek(0)
    xlsx_bytes = output.read()
    if xlsx_bytes[:4] != b"PK\x03\x04":
        from openpyxl import Workbook
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "오류"
        ws2["A1"] = "Excel 파일 직렬화 실패 (빈 출력)"
        fb = BytesIO()
        wb2.save(fb)
        fb.seek(0)
        return fb.read()

    return xlsx_bytes


# ── CSS ──────────────────────────────────────────────────────────────────────

def inject_css():
    st.markdown("""
    <style>
    /* ── 기본 리셋: 햄버거·푸터·툴바 숨김 ── */
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }
    [data-testid="stToolbar"] { visibility: hidden; }
    [data-testid="stHeader"] { background: transparent; }
    /* 사이드바 완전 숨김 (expander로 대체) */
    section[data-testid="stSidebar"],
    [data-testid="collapsedControl"],
    [data-testid="stSidebarCollapsedControl"] {
        display: none !important;
    }

    /* ── 전체 컨테이너: 좌우 대칭 여백 + 고정 너비 ── */
    .block-container {
        padding: 1rem 3rem 5rem !important;
        max-width: 720px !important;
    }

    /* ── 앱 헤더 ── */
    .app-header {
        padding: 1.5rem 0 1.25rem;
        border-bottom: 2px solid #E2E8F0;
        margin-bottom: 2rem;
    }
    .app-header h1 {
        font-size: 1.75rem;
        font-weight: 800;
        color: #0F172A;
        margin: 0 0 0.3rem;
        letter-spacing: -0.03em;
    }
    .app-header p {
        font-size: 0.85rem;
        color: #94A3B8;
        margin: 0;
    }

    /* ── 검색 카드 ── */
    .card {
        background: #F8FAFC;
        border: 1px solid #E2E8F0;
        border-radius: 12px;
        padding: 1.5rem 1.5rem 0.5rem;
        margin-bottom: 1.5rem;
    }
    .card-title {
        font-size: 0.68rem;
        font-weight: 700;
        letter-spacing: 0.1em;
        text-transform: uppercase;
        color: #94A3B8;
        margin-bottom: 0.85rem;
    }

    /* ── 기업 정보 배지 ── */
    .corp-badge {
        display: flex;
        align-items: center;
        gap: 0.5rem;
        background: #EFF6FF;
        border: 1px solid #BFDBFE;
        border-radius: 10px;
        padding: 0.65rem 1.1rem;
        margin: 0.5rem 0 1rem;
    }
    .corp-name { font-size: 1rem; font-weight: 700; color: #1E40AF; }
    .corp-meta { font-size: 0.8rem; color: #3B82F6; }
    .corp-tag {
        font-size: 0.68rem; font-weight: 700;
        background: #DBEAFE; color: #1D4ED8;
        padding: 0.15rem 0.5rem; border-radius: 4px;
    }

    /* ── 다운로드 버튼 ── */
    [data-testid="stDownloadButton"] button {
        background: #2563EB !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
        padding: 0.65rem 1.5rem !important;
        transition: background 0.15s !important;
        width: 100% !important;
    }
    [data-testid="stDownloadButton"] button:hover {
        background: #1D4ED8 !important;
    }

    /* ── 제출 버튼 ── */
    [data-testid="stFormSubmitButton"] button {
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 0.95rem !important;
    }


    /* ── input / selectbox ── */
    [data-baseweb="input"] > div,
    [data-baseweb="select"] > div:first-child {
        border-radius: 8px !important;
    }

    /* ── 구분선 ── */
    hr { border: none; border-top: 1px solid #E2E8F0; margin: 1.25rem 0; }
    </style>
    """, unsafe_allow_html=True)


# ── UI ───────────────────────────────────────────────────────────────────────

st.set_page_config(page_title="DARTSHEET", page_icon=None, layout="centered",
                   initial_sidebar_state="collapsed")
inject_css()

# ── 인증 상태 확인 ────────────────────────────────────────────────────────────
_api_ready = (
    st.session_state.get("_master_auth")
    or bool(st.session_state.get("_user_api_key", "").strip())
)

# ── 로그인 페이지 ─────────────────────────────────────────────────────────────
if not _api_ready:
    st.markdown("""
    <div style="text-align:center; padding: 3rem 0 2rem;">
      <div style="font-size:2rem; font-weight:900; letter-spacing:-0.04em; color:#0F172A;">
        DARTSHEET
      </div>
      <div style="font-size:0.85rem; color:#94A3B8; margin-top:0.4rem;">
        국내 상장사 재무제표 자동 수집
      </div>
    </div>
    """, unsafe_allow_html=True)

    # 카드 컨테이너
    with st.container():
        # 탭: 마스터 코드 / 개인 API 키
        _tab_personal, _tab_master = st.tabs(["개인 API 키", "관리자 코드"])

        with _tab_master:
            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            _code_input = st.text_input(
                "마스터 코드",
                max_chars=4,
                type="password",
                placeholder="4자리 코드 입력",
                key="_login_code",
            )
            if st.button("입장하기", use_container_width=True, type="primary", key="_btn_master"):
                if _code_input and _MASTER_CODE and _code_input == _MASTER_CODE:
                    st.session_state["_master_auth"] = True
                    st.rerun()
                else:
                    st.error("코드가 올바르지 않습니다.")

        with _tab_personal:
            st.markdown("<div style='height:0.5rem'></div>", unsafe_allow_html=True)
            st.markdown(
                "<small style='color:#94A3B8'>키 발급: "
                "<a href='https://opendart.fss.or.kr/uat/uia/easyLogin.do' target='_blank'>"
                "opendart.fss.or.kr</a></small>",
                unsafe_allow_html=True,
            )
            _key_input = st.text_input(
                "DART API 키",
                type="password",
                placeholder="40자리 API 키 입력",
                key="_login_key",
            )
            if st.button("입장하기", use_container_width=True, type="primary", key="_btn_personal"):
                if len(_key_input.strip()) >= 30:
                    st.session_state["_user_api_key"] = _key_input.strip()
                    st.rerun()
                else:
                    st.error("유효한 API 키를 입력해주세요 (40자리).")

    st.markdown("""
    <div style="text-align:center; margin-top:2.5rem; color:#CBD5E1; font-size:0.78rem;">
        Contact: 010-9688-3575
    </div>
    """, unsafe_allow_html=True)

    st.stop()

# ── 여기서부터 인증 완료된 사용자만 접근 ─────────────────────────────────────

_hcol1, _hcol2 = st.columns([5, 1])
with _hcol1:
    st.markdown("""
    <div class="app-header">
      <h1>DARTSHEET</h1>
      <p>국내 상장사의 재무제표를 DART에서 자동으로 가져와 연도별 엑셀로 정리합니다.</p>
    </div>
    """, unsafe_allow_html=True)
with _hcol2:
    st.markdown("""
    <div style="padding-top:1.7rem; text-align:right;">
      <a href="." style="
        background:#EF4444; color:white; text-decoration:none;
        border-radius:5px; font-size:0.6rem; font-weight:600;
        padding:0.18rem 0.45rem; white-space:nowrap; line-height:1;
        display:inline-block;
      ">로그아웃</a>
    </div>
    """, unsafe_allow_html=True)

# 기업코드 목록 로드
with st.spinner("DART 기업 목록 초기화 중..."):
    try:
        corp_df = load_corp_codes()
    except Exception as e:
        st.error(f"DART 기업 목록 로드 실패: {e}")
        st.stop()

# 검색 폼
with st.form("main_form"):
    query = st.text_input(
        "회사명 또는 종목코드",
        placeholder="예: 삼성전자 / 005930",
        label_visibility="collapsed",
    )

    col1, col2, col3 = st.columns([2, 2, 3])
    today = datetime.date.today()
    default_year = today.year - 1

    with col1:
        end_year = st.number_input("기준 연도", min_value=2010,
                                   max_value=today.year, value=default_year, step=1)
    with col2:
        num_years = st.number_input("수집 연수", min_value=1, max_value=10, value=5, step=1)
    with col3:
        report_type = st.selectbox("보고서 종류", list(REPORT_TYPES.keys()))

    submitted = st.form_submit_button("검색 및 수집", use_container_width=True, type="primary")

# ── 검색 & Excel 생성 ────────────────────────────────────────────────────────
# [구조]
# 1) 폼 제출(submitted=True) 시 → 검색 결과·파라미터를 session_state에 저장
# 2) selectbox 선택 rerun 포함 모든 rerun에서 → session_state 기반으로 UI 표시 & Excel 빌드
# → selectbox 선택 시 submitted=False가 되어도 build_excel이 올바르게 호출됨

if submitted and query.strip():
    results = search_company(query.strip(), corp_df)
    if results.empty:
        st.error("검색 결과가 없습니다. 회사명을 다시 확인해주세요.")
        st.session_state.pop("_search_state", None)
        st.stop()
    years      = list(range(int(end_year) - int(num_years) + 1, int(end_year) + 1))
    reprt_code = REPORT_TYPES[report_type]
    # 결과·파라미터 저장 — selectbox rerun 시에도 유지됨
    st.session_state["_search_state"] = {
        "results":     results.head(10).to_dict("records"),
        "years":       years,
        "reprt_code":  reprt_code,
        "report_type": report_type,
        "num_years":   num_years,
    }
    # 새 검색 시 이전 Excel 파일 초기화
    st.session_state.pop("_excel_bytes", None)
    st.session_state.pop("_built_key",   None)

elif submitted:
    st.warning("회사명 또는 종목코드를 입력해주세요.")

# ── 검색 결과 UI + Excel 빌드 (session_state 기반, 모든 rerun에서 실행) ──────
if "_search_state" in st.session_state:
    _ss        = st.session_state["_search_state"]
    _results   = pd.DataFrame(_ss["results"])
    _years     = _ss["years"]
    _reprt     = _ss["reprt_code"]
    _rtype_lbl = _ss["report_type"]
    _n_yr      = _ss["num_years"]

    if len(_results) > 1:
        _options = [
            f"{r['corp_name']}  ·  {r['stock_code'] or '비상장'}"
            for r in _ss["results"]
        ]
        _choice = st.selectbox("검색된 기업 선택", range(len(_options)),
                               format_func=lambda i: _options[i],
                               key="_corp_choice")
        _sel = _results.iloc[_choice]
    else:
        _sel = _results.iloc[0]

    _corp_code  = _sel["corp_code"]
    _corp_name  = _sel["corp_name"]
    _stock_code = _sel["stock_code"]
    _year_range = f"{_years[0]} – {_years[-1]}"

    st.markdown(f"""
    <div class="corp-badge">
      <span class="corp-name">{_corp_name}</span>
      <span class="corp-tag">{_stock_code or "비상장"}</span>
      <span class="corp-meta">· {_year_range} · {_rtype_lbl.split()[0]}</span>
    </div>
    """, unsafe_allow_html=True)

    # _excel_bytes 없을 때만 재생성
    # (새 검색 시 submitted 블록에서 pop → 여기서 재빌드 / 다운로드 rerun은 bytes 있으므로 스킵)
    if "_excel_bytes" not in st.session_state:
        try:
            _excel_bytes = build_excel(_corp_code, _corp_name, _stock_code, _years, _reprt)
            st.session_state["_excel_bytes"]    = _excel_bytes
            st.session_state["_excel_filename"] = (
                f"{_corp_name}_재무제표_{_years[0]}-{_years[-1]}.xlsx"
            )
            st.session_state["_excel_meta"] = (_year_range, _n_yr, _rtype_lbl.split()[0])
        except Exception as _build_err:
            st.error(f"Excel 생성 오류: {_build_err}")
            st.exception(_build_err)
            st.session_state.pop("_excel_bytes", None)

# ── 다운로드 버튼 (항상 렌더링 — session_state 기반) ────────────────────────
_dl_bytes = st.session_state.get("_excel_bytes")
if _dl_bytes:
    _yr_range, _n_yr, _rtype = st.session_state.get(
        "_excel_meta", ("–", "–", "–")
    )
    st.markdown(f"""
    <div style="display:flex;gap:12px;margin:10px 0;">
      <span style="background:#EFF6FF;color:#1E3A5F;padding:8px 20px;border-radius:8px;font-size:15px;font-weight:600;">
        📅 {_yr_range}
      </span>
      <span style="background:#EFF6FF;color:#1E3A5F;padding:8px 20px;border-radius:8px;font-size:15px;font-weight:600;">
        {_n_yr}개년
      </span>
      <span style="background:#EFF6FF;color:#1E3A5F;padding:8px 20px;border-radius:8px;font-size:15px;font-weight:600;">
        {_rtype}
      </span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<hr>", unsafe_allow_html=True)
    st.download_button(
        label="엑셀 다운로드  (.xlsx)",
        data=_dl_bytes,
        file_name=st.session_state.get("_excel_filename", "재무제표.xlsx"),
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
